import openpyxl
from openpyxl import Workbook
import os

# Создаем новую книгу
wb = Workbook()
ws = wb.active
ws.title = "Haulhwr"

# Записываем заголовки
ws['A1'] = "Заголовок 1"
ws['B1'] = "Заголовок 2"
ws['C1'] = "Заголовок 3"

# Добавляем данные
data = [
    ["Зов", "Зовчик", 24],
    ["Какавоз", "Какавозик", 32],
    ["ГАН", "Ганович", 44]
]

for row in data:
    ws.append(row)

# Получаем путь к папке, где находится этот скрипт
script_dir = os.path.dirname(os.path.abspath(__file__))
# Создаем полный путь для сохранения
file_path = os.path.join(script_dir, 'example.xlsx')

# Сохраняем
wb.save(file_path)

# Выводим содержимое файла
print(f"Файл сохранен: {file_path}")
print("Содержимое файла example.xlsx:")
for row in ws.iter_rows(values_only=True):
    print("\t".join(str(cell) if cell is not None else "" for cell in row))